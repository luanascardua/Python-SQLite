from openpyxl import load_workbook
import time

class Sheet():


    def __init__(self, wb = 'insertData.xlsx'):

        global sheets
        #self.firstLine = int
        self.wb = load_workbook(wb)
        sheets = self.wb.sheetnames
    
        print(f'Sheet: {sheets}')


    def countLines(self, sheet):

        global ws

        ws = self.wb[sheet]

        self.lastLine = len(ws['B']) + 1
        data =  [
            x.value for x in ws['I']
            if x.value != None
        ]      
        self.firstLine = len(data) + 1
        print(f'1 linha: {self.firstLine}\núltima: {self.lastLine}')


    def getDataUser(self, lines):
    
        self.name    = ws.cell(row=lines, column=2).value
        self.sigla   = ws.cell(row=lines, column=3).value
        self.idade   = ws.cell(row=lines, column=4).value

        
    def getDataTipoAvaliacao(self, lines):

        self.sigla     = ws.cell(row=lines, column=1).value
        self.descricao = ws.cell(row=lines, column=2).value


    def getDataAvaliacao(self, lines):

        self.sigla          = ws.cell(row=lines, column=2).value
        self.descricao      = ws.cell(row=lines, column=3).value
        self.tipoAvaliacao  = ws.cell(row=lines, column=4).value
        self.dataInicio     = ws.cell(row=lines, column=5).value
        self.dataTermino    = ws.cell(row=lines, column=6).value
        self.notaAprovacao  = ws.cell(row=lines, column=7).value
        self.idUser         = ws.cell(row=lines, column=8).value


    def getDataAvaliacaoAluno(self, lines):

        self.idUser         = ws.cell(row=lines, column=1).value
        self.idAvaliacao    = ws.cell(row=lines, column=2).value
        self.dataConclusao  = ws.cell(row=lines, column=3).value
        self.nota           = ws.cell(row=lines, column=4).value
        self.situacao       = ws.cell(row=lines, column=5).value



executeWb = Sheet()
executeWb.countLines(sheets[0])


#for lines in (firstLine, lastLine):
'''functions = locals()
print(f'Funções: {functions}')
funcao = 'getData'
functions[funcao]()'''